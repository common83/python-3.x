#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/6/21 19:06
# @Author  : Kevindi
# @Site    : 
# @File    : get_data_from_zhidao.py.py
# @Software: PyCharm Community Edition
# @Desc    :
import re
from urllib import parse
import openpyxl

import requests
import time

from  baidu_zhidao.proxy_tencent import proxies
from bs4 import BeautifulSoup

file_name = './球员覆盖 - 副本.xlsx'
sheet_name = 'Sheet1'
# 读写excel表格
work_file = openpyxl.load_workbook(file_name)
work_sheet = work_file[sheet_name]


# 读取所有的查询语料
def get_query():
    query_list = []
    i = 0
    col_index = 0
    for row in work_sheet.rows:
        if i == 0:
            for cell in row:
                if cell.value != 'strText':
                    col_index += 1
                else:
                    print("find:",cell.value)
                    break  # 找到列
                # print(cell.value)
            i += 1
        else:
            # print(list(row)[col_index].value)
            query_list.append(list(row)[col_index].value)

    print('value count=', len(query_list))
    return query_list


# 暴力删除不可编码为 GBK的字符
def killAnUnseen(s):
    try:
        s.encode('gbk')
    except UnicodeEncodeError as err:
        info = str(err)
        st = re.search('\\\\u[a-f0-9]{4}|\\\\x[a-f0-9]{2}', info).group()[2:]
        x = int(st, 16)
        return s.replace(chr(x), "")
    return s


# 从百科网站获取结果
def get_baike_value(text_t):
    # text_t = '俄罗斯队的主教练'
    text_t = killAnUnseen(text_t) # url需要进行GBK编码
    url = 'https://zhidao.baidu.com/search?lm=0&rn=10&pn=0&fr=search&ie=gbk&word='
    url += parse.quote(text_t.strip(), encoding='GBK')
    # print('url=', url)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.26 Safari/537.36 Core/1.63.5383.400 QQBrowser/10.0.1313.400'
    }
    r = requests.request('GET', url=url,headers=headers, proxies=proxies)
    # requests.models.Response

    # print(type(r))
    # print(r.content)
    soup = BeautifulSoup(r.content, 'lxml')
    # print(soup)
    data_store = dict()
    data_list = list()
    data_store['url'] = url
    data_store['query'] = soup.find('input',attrs={'class': 'hdi', 'id': 'kw', 'name': 'word'}).get('value')
    dd_text = ""
    try:
        dd_text = soup.find('dd', attrs={'class': 'dd answer'}).text
        # print("dlt", dd_list)
        data_store['answer'] = dd_text
        print('ok ', data_store)
    except:
        print('error: ', data_store)
        # print(soup)
    return dd_text

max_col = 12
c = work_sheet.cell(row=1, column=max_col + 1)
c.value = "baike_answer"

query_list = get_query()
row_id = 2
MAX_COUNT  = 100
i = 0
try:
    for text in query_list:
        baike_answer = get_baike_value(text)
        c = work_sheet.cell(row=row_id, column=max_col+1)
        # c.value = baike_answer.strip('\n')
        c.value = baike_answer.strip()
        row_id += 1
        i += 1
        if i%MAX_COUNT == 0:
            time.sleep(10)
finally:
    work_file.save(file_name)
